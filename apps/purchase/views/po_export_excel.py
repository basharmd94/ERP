from django.http import HttpResponse
from django.db import connection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def po_export_excel(request, grn_number):
    zid = request.session.get('current_zid')
    if not zid:
        return HttpResponse("Session ZID not found", status=401)

    header_sql = """
        SELECT
            grn.xgrnnum,
            grn.xdate,
            grn.xsup,
            s.xorg AS supplier_name,
            s.xadd1 AS supplier_address,
            grn.xproj,
            grn.xwh,
            grn.xrem,
            grn.xdtwotax,
            grn.xdiscamt,
            grn.xdttax,
            grn.xtotamt,
            grn.xpornum,
            po.xdate AS po_date,
            grn.xconfirmt,
            grn.zemail,
            grn.xstatusgrn,
            grn.xref
        FROM pogrn grn
        LEFT JOIN casup s ON s.zid = grn.zid AND s.xsup = grn.xsup
        LEFT JOIN poord po ON po.zid = grn.zid AND po.xpornum = grn.xpornum
        WHERE grn.zid = %s AND grn.xgrnnum = %s
        LIMIT 1
    """

    details_sql = """
        SELECT
            d.xrow,
            d.xitem,
            i.xdesc,
            i.xunitstk,
            d.xqty,
            d.xqtygrn,
            d.xrate,
            d.xlineamt
        FROM pogdt d
        LEFT JOIN caitem i ON i.zid = d.zid AND i.xitem = d.xitem
        WHERE d.zid = %s AND d.xgrnnum = %s
        ORDER BY d.xrow
    """

    with connection.cursor() as cursor:
        cursor.execute(header_sql, [zid, grn_number])
        hr = cursor.fetchone()
        if not hr:
            return HttpResponse(f"GRN not found: {grn_number}", status=404)

        header = {
            'xgrnnum': hr[0] or '',
            'xdate': hr[1],
            'xsup': hr[2] or '',
            'supplier_name': hr[3] or '',
            'supplier_address': hr[4] or '',
            'xproj': hr[5] or '',
            'xwh': hr[6] or '',
            'xrem': hr[7] or '',
            'xdtwotax': float(hr[8]) if hr[8] is not None else 0.0,
            'xdiscamt': float(hr[9]) if hr[9] is not None else 0.0,
            'xdttax': float(hr[10]) if hr[10] is not None else 0.0,
            'xtotamt': float(hr[11]) if hr[11] is not None else 0.0,
            'xpornum': hr[12] or '',
            'po_date': hr[13],
            'xconfirmt': hr[14],
            'zemail': hr[15] or '',
            'xstatusgrn': hr[16] or '',
            'xref': hr[17] or ''
        }

        cursor.execute(details_sql, [zid, grn_number])
        rows = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "GRN"

    title_font = Font(name='Arial', size=16, bold=True)
    header_font = Font(name='Arial', size=12, bold=True)
    data_font = Font(name='Arial', size=10)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    right = Alignment(horizontal='right', vertical='center')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

    ws.merge_cells('A1:H1')
    ws['A1'] = 'GOODS RECEIPT NOTE'
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    row = 3
    ws[f'A{row}'] = 'MRR No:'
    ws[f'B{row}'] = header.get('xgrnnum', '-')
    ws[f'D{row}'] = 'MRR Date:'
    ws[f'E{row}'] = header.get('xdate').strftime('%Y-%m-%d') if header.get('xdate') else ''

    row += 1
    ws[f'A{row}'] = "Supplier's ID:"
    ws[f'B{row}'] = header.get('xsup', '-')
    ws[f'D{row}'] = 'Supplier Name:'
    ws[f'E{row}'] = header.get('supplier_name', '-')

    row += 1
    ws[f'A{row}'] = 'Purchase Order No:'
    ws[f'B{row}'] = header.get('xpornum', '-')
    ws[f'D{row}'] = 'PO Date:'
    ws[f'E{row}'] = header.get('po_date').strftime('%Y-%m-%d') if header.get('po_date') else ''

    row += 1
    ws[f'A{row}'] = 'Project:'
    ws[f'B{row}'] = header.get('xproj', '-')
    ws[f'D{row}'] = 'Warehouse:'
    ws[f'E{row}'] = header.get('xwh', '-')

    row += 1
    ws[f'A{row}'] = 'Confirmed Date:'
    ws[f'B{row}'] = header.get('xconfirmt').strftime('%Y-%m-%d') if header.get('xconfirmt') else ''
    ws[f'D{row}'] = 'GRN Status:'
    ws[f'E{row}'] = header.get('xstatusgrn', '-')

    row += 1
    ws[f'A{row}'] = 'INVOICE#:'
    ws[f'B{row}'] = header.get('xref', '-')
    ws[f'D{row}'] = 'Prepared By:'
    ws[f'E{row}'] = header.get('zemail', '-')

    row += 3
    headers = ['Sl No', 'Item Code', 'Description of Goods', 'Unit', 'Quantity Supplied', 'Quantity Received', 'Unit Value', 'Total Value']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.alignment = center
        c.border = thin
        c.fill = fill

    for idx, r in enumerate(rows, start=1):
        row += 1
        ws.cell(row=row, column=1, value=idx).alignment = center
        ws.cell(row=row, column=2, value=r[1] or '').alignment = left
        ws.cell(row=row, column=3, value=(r[2] or (r[1] or ''))).alignment = left
        ws.cell(row=row, column=4, value=r[3] or '').alignment = center
        ws.cell(row=row, column=5, value=float(r[4]) if r[4] is not None else 0.0).alignment = right
        ws.cell(row=row, column=6, value=float(r[5]) if r[5] is not None else 0.0).alignment = right
        ws.cell(row=row, column=7, value=float(r[6]) if r[6] is not None else 0.0).alignment = right
        ws.cell(row=row, column=8, value=float(r[7]) if r[7] is not None else 0.0).alignment = right
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin
            ws.cell(row=row, column=col).font = data_font

    row += 2
    ws[f'A{row}'] = 'Sub Total'
    ws[f'B{row}'] = header.get('xdtwotax', 0)
    row += 1
    ws[f'A{row}'] = 'Discount'
    ws[f'B{row}'] = header.get('xdiscamt', 0)
    row += 1
    ws[f'A{row}'] = 'Tax'
    ws[f'B{row}'] = header.get('xdttax', 0)
    row += 1
    ws[f'A{row}'] = 'Total'
    ws[f'B{row}'] = header.get('xtotamt', 0)

    for column in ws.columns:
        max_len = 0
        letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                v = str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="GRN_{grn_number}.xlsx"'
    wb.save(response)
    return response
