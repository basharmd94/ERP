from django.http import HttpResponse
from django.db import connection
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from xhtml2pdf import pisa
from io import BytesIO, StringIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib import messages
import csv


@login_required
def daily_sales_report_export(request):
    # Get current ZID from session
    session_zid = request.session.get('current_zid')

    # Get parameters from request
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    report_format = request.GET.get('report_type', 'pdf')

    # Debug: Log the received parameters
    print(f"DEBUG: from_date={from_date}, to_date={to_date}, report_format={report_format}")
    print(f"DEBUG: All GET parameters: {dict(request.GET)}")

    if not from_date or not to_date:
        return HttpResponse('From Date and To Date are required', status=400)

    # SQL query for daily sales report
    daily_sales_sql = """
        SELECT
            xdate,
            xordernum,
            xsalescat,
            xdtcomm,
            (xtotamt) - (xdtcomm + xdiscf + xdtdisc) as cash_amount,
            xdiscf + xdtdisc as discount,
            xtotamt
        FROM opordview
        WHERE zid = %s
          AND xdate BETWEEN %s AND %s
        ORDER BY xdate, xordernum
    """

    with connection.cursor() as cursor:
        # Get daily sales data
        cursor.execute(daily_sales_sql, [session_zid, from_date, to_date])
        rows = cursor.fetchall()

        # Get business information from session (populated by BusinessInfoMiddleware)
        business_info = request.session.get('business_info', {})
        business_data = {
            'business_id': business_info.get('zid'),
            'business_name': business_info.get('business_name', ''),
            'business_address': business_info.get('business_address', ''),
            'business_mobile': business_info.get('business_mobile', ''),
            'business_email': business_info.get('business_email', ''),
            'business_website': business_info.get('business_website', ''),
        }

        # Process the sales data
        sales_data = []
        for row in rows:
            sales_item = {
                'xdate': row[0].strftime('%Y-%m-%d') if row[0] else '',
                'xordernum': row[1] or '',
                'xsalescat': row[2] or '',  # Bank
                'xdtcomm': float(row[3]) if row[3] else 0.0,  # Card Amount
                'cash_amount': float(row[4]) if row[4] else 0.0,  # Cash Amount
                'discount': float(row[5]) if row[5] else 0.0,  # Discount
                'xtotamt': float(row[6]) if row[6] else 0.0  # Total
            }
            sales_data.append(sales_item)

    # Calculate grand totals
    grand_totals = {
        'total_card_amount': sum(item['xdtcomm'] for item in sales_data),
        'total_cash_amount': sum(item['cash_amount'] for item in sales_data),
        'total_discount': sum(item['discount'] for item in sales_data),
        'grand_total': sum(item['xtotamt'] for item in sales_data)
    }

    # Check date range for PDF export - automatically switch to CSV if >= 31 days
    if report_format.lower() == 'pdf':
        try:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d')
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d')
            date_diff = (to_date_obj - from_date_obj).days + 1  # +1 to include both dates

            if date_diff >= 7:
                print(f"DEBUG: Date range is {date_diff} days (>=7), automatically switching PDF to CSV for performance")
                messages.warning(request, f"Date range is {date_diff} days (>=7), automatically switching to CSV format for performance.")
                report_format = 'csv'  # Override format to CSV
        except ValueError:
            # If date parsing fails, continue with original format
            print("DEBUG: Date parsing failed, continuing with original format")
            pass

    # Generate report based on format
    print(f"DEBUG: Checking format - report_format.lower() = '{report_format.lower()}'")
    if report_format.lower() == 'excel':
        print("DEBUG: Generating Excel report")
        return generate_excel_report(sales_data, business_data, from_date, to_date)
    elif report_format.lower() == 'csv':
        print("DEBUG: Generating CSV report")
        return generate_csv_report(sales_data, business_data, from_date, to_date)
    else:
        print("DEBUG: Generating PDF report")
        return generate_pdf_report(request, sales_data, business_data, from_date, to_date, grand_totals)


def generate_pdf_report(request, sales_data, business_data, from_date, to_date, grand_totals):
    """Generate PDF report using xhtml2pdf"""

    # Render HTML template
    html_string = render_to_string('reports/daily_sales_report_export.html', {
        'business': business_data,
        'sales_data': sales_data,
        'from_date': from_date,
        'to_date': to_date,
        'print_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'total_records': len(sales_data),
        'grand_totals': grand_totals,
        'report_title': 'Daily Sales Report',
    }, request=request)

    # Generate PDF using xhtml2pdf
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)

    if not pdf.err:
        # Return PDF response
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="daily_sales_report_{from_date}_to_{to_date}.pdf"'
        return response
    else:
        return HttpResponse('Error generating PDF', status=500)


def generate_excel_report(sales_data, business_data, from_date, to_date):
    """Generate Excel report using openpyxl"""

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Sales Report"

    # Define styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Add business header
    ws.merge_cells('A1:G1')
    ws['A1'] = business_data['business_name']
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment

    ws.merge_cells('A2:G2')
    ws['A2'] = business_data['business_address']
    ws['A2'].alignment = center_alignment

    ws.merge_cells('A3:G3')
    ws['A3'] = f"Mobile: {business_data['business_mobile']} | Email: {business_data['business_email']}"
    ws['A3'].alignment = center_alignment

    ws.merge_cells('A4:G4')
    ws['A4'] = f"Website: {business_data['business_website']}"
    ws['A4'].alignment = center_alignment

    # Add report title
    ws.merge_cells('A6:G6')
    ws['A6'] = "DAILY SALES REPORT"
    ws['A6'].font = title_font
    ws['A6'].alignment = center_alignment

    # Add date range
    ws.merge_cells('A7:G7')
    ws['A7'] = f"From: {from_date} To: {to_date}"
    ws['A7'].alignment = center_alignment

    # Add column headers
    headers = ['Date', 'Order Number', 'Bank', 'Card Amount', 'Cash Amount', 'Discount', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment

    # Add data rows
    for row_idx, item in enumerate(sales_data, 10):
        ws.cell(row=row_idx, column=1, value=item['xdate']).border = border
        ws.cell(row=row_idx, column=2, value=item['xordernum']).border = border
        ws.cell(row=row_idx, column=3, value=item['xsalescat']).border = border
        ws.cell(row=row_idx, column=4, value=item['xdtcomm']).border = border
        ws.cell(row=row_idx, column=5, value=item['cash_amount']).border = border
        ws.cell(row=row_idx, column=6, value=item['discount']).border = border
        ws.cell(row=row_idx, column=7, value=item['xtotamt']).border = border

    # Auto-adjust column widths
    for col in range(1, 8):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return Excel response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="daily_sales_report_{from_date}_to_{to_date}.xlsx"'

    return response


def generate_csv_report(sales_data, business_data, from_date, to_date):
    """Generate CSV report using Python's csv module"""

    # Create StringIO buffer
    output = StringIO()
    writer = csv.writer(output)

    # Write business header information
    writer.writerow([business_data['business_name']])
    writer.writerow([business_data['business_address']])
    writer.writerow([f"Mobile: {business_data['business_mobile']} | Email: {business_data['business_email']}"])
    writer.writerow([f"Website: {business_data['business_website']}"])
    writer.writerow([])  # Empty row

    # Write report title and date range
    writer.writerow(['DAILY SALES REPORT'])
    writer.writerow([f'From: {from_date} To: {to_date}'])
    writer.writerow([f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    writer.writerow([])  # Empty row

    # Write column headers
    headers = ['Date', 'Order Number', 'Bank', 'Card Amount', 'Cash Amount', 'Discount', 'Total']
    writer.writerow(headers)

    # Write data rows
    for item in sales_data:
        writer.writerow([
            item['xdate'],
            item['xordernum'],
            item['xsalescat'],
            item['xdtcomm'],
            item['cash_amount'],
            item['discount'],
            item['xtotamt']
        ])

    # Write summary totals
    writer.writerow([])  # Empty row
    writer.writerow(['SUMMARY'])
    writer.writerow(['Total Records:', len(sales_data)])
    writer.writerow(['Total Card Amount:', sum(item['xdtcomm'] for item in sales_data)])
    writer.writerow(['Total Cash Amount:', sum(item['cash_amount'] for item in sales_data)])
    writer.writerow(['Total Discount:', sum(item['discount'] for item in sales_data)])
    writer.writerow(['Grand Total:', sum(item['xtotamt'] for item in sales_data)])

    # Get CSV content
    csv_content = output.getvalue()
    output.close()

    # Create HTTP response
    response = HttpResponse(csv_content, content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="daily_sales_report_{from_date}_to_{to_date}.csv"'

    return response
