"""
Sales Return Excel Export View
Generates Excel file for sales return data
"""

from django.http import HttpResponse
from django.views import View
from django.db import connection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class SalesReturnExcelExportView(View):
    """
    Export sales return data to Excel format
    """

    def get(self, request, transaction_id):
        """
        Generate and return Excel file for sales return
        """
        try:
            # Get sales return data
            header_data, line_items, totals = self._get_sales_return_data(transaction_id)

            if not header_data:
                return HttpResponse("Sales return not found", status=404)

            # Create Excel workbook
            workbook = self._create_excel_workbook(header_data, line_items, totals)

            # Prepare response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{transaction_id}.xlsx"'

            # Save workbook to response
            workbook.save(response)
            return response

        except Exception as e:
            return HttpResponse(f"Error generating Excel file: {str(e)}", status=500)

    def _get_sales_return_data(self, transaction_id):
        """
        Fetch sales return data from database using the same query structure as sales_return_detail.py
        """
        current_zid = self.request.session.get('current_zid')
        if not current_zid:
            return None, [], {}

        with connection.cursor() as cursor:
            # Use the same SQL query as sales_return_detail.py
            query = """
                SELECT
                    t.ximtmptrn,
                    t.xdate,
                    t.xrem,
                    t.xproj,
                    t.xwh,
                    t.zemail,
                    t.xglref,
                    t.xstatustrn,
                    d.xtorlno,
                    d.xitem,
                    d.xqtyord,
                    d.xunit,
                    d.ximtrnnum,
                    d.xrate,
                    d.xval,
                    d.xlineamt
                FROM imtemptrn t
                INNER JOIN imtemptdt d
                    ON t.ximtmptrn = d.ximtmptrn
                    AND t.zid = d.zid
                WHERE t.zid = %s
                  AND t.ximtmptrn = %s
            """

            cursor.execute(query, [current_zid, transaction_id])
            rows = cursor.fetchall()

            if not rows:
                return None, [], {}

            # Process the data
            header_data = None
            line_items = []

            for row in rows:
                # Header data (same for all rows)
                if header_data is None:
                    header_data = {
                        'ximtmptrn': row[0],
                        'xdate': row[1].strftime('%Y-%m-%d') if row[1] else '',
                        'xrem': row[2] or '',
                        'xproj': row[3] or '',
                        'xwh': row[4] or '',
                        'zemail': row[5] or '',
                        'xglref': row[6] or '',
                        'xstatustrn': row[7] or ''
                    }

                # Line item data
                line_item = {
                    'xtorlno': row[8] or '',
                    'xitem': row[9] or '',
                    'xqtyord': float(row[10]) if row[10] else 0.0,
                    'xunit': row[11] or '',
                    'ximtrnnum': row[12] or '',
                    'xrate': float(row[13]) if row[13] else 0.0,
                    'xval': float(row[14]) if row[14] else 0.0,
                    'xlineamt': float(row[15]) if row[15] else 0.0
                }
                line_items.append(line_item)

            # Calculate totals
            totals = {
                'item_count': len(line_items),
                'total_quantity': sum(item['xqtyord'] for item in line_items),
                'total_value': sum(item['xval'] for item in line_items),
                'total_line_amount': sum(item['xlineamt'] for item in line_items),
            }

            return header_data, line_items, totals

    def _create_excel_workbook(self, header_data, line_items, totals):
        """
        Create formatted Excel workbook
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sales Return"

        # Define styles
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=12, bold=True)
        data_font = Font(name='Arial', size=10)

        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

        # Title
        worksheet.merge_cells('A1:H1')
        worksheet['A1'] = 'SALES RETURN REPORT'
        worksheet['A1'].font = title_font
        worksheet['A1'].alignment = center_alignment

        # Header information
        row = 3
        worksheet[f'A{row}'] = 'Transaction ID:'
        worksheet[f'B{row}'] = header_data.get('ximtmptrn', '-')
        worksheet[f'D{row}'] = 'Date:'
        worksheet[f'E{row}'] = header_data.get('xdate', '-')

        row += 1
        worksheet[f'A{row}'] = 'Project:'
        worksheet[f'B{row}'] = header_data.get('xproj', '-')
        worksheet[f'D{row}'] = 'Warehouse:'
        worksheet[f'E{row}'] = header_data.get('xwh', '-')

        row += 1
        worksheet[f'A{row}'] = 'Status:'
        worksheet[f'B{row}'] = header_data.get('xstatustrn', '-')
        worksheet[f'D{row}'] = 'GL Reference:'
        worksheet[f'E{row}'] = header_data.get('xglref', '-')

        row += 1
        worksheet[f'A{row}'] = 'Email:'
        worksheet[f'B{row}'] = header_data.get('zemail', '-')
        worksheet[f'D{row}'] = 'Remarks:'
        worksheet[f'E{row}'] = header_data.get('xrem', '-')

        # Line items table
        row += 3
        headers = ['Line No.', 'Item Code', 'Quantity', 'Unit', 'Rate', 'Value', 'Line Amount', 'Transaction No.']

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = header_fill

        # Line items data
        for item in line_items:
            row += 1
            worksheet.cell(row=row, column=1, value=item.get('xtorlno', '-')).alignment = center_alignment
            worksheet.cell(row=row, column=2, value=item.get('xitem', '-')).alignment = left_alignment
            worksheet.cell(row=row, column=3, value=item.get('xqtyord', 0)).alignment = right_alignment
            worksheet.cell(row=row, column=4, value=item.get('xunit', '-')).alignment = center_alignment
            worksheet.cell(row=row, column=5, value=item.get('xrate', 0)).alignment = right_alignment
            worksheet.cell(row=row, column=6, value=item.get('xval', 0)).alignment = right_alignment
            worksheet.cell(row=row, column=7, value=item.get('xlineamt', 0)).alignment = right_alignment
            worksheet.cell(row=row, column=8, value=item.get('ximtrnnum', '-')).alignment = left_alignment

            # Apply borders
            for col in range(1, 9):
                worksheet.cell(row=row, column=col).border = thin_border
                worksheet.cell(row=row, column=col).font = data_font

        # Totals section
        row += 2
        worksheet[f'A{row}'] = 'SUMMARY TOTALS'
        worksheet[f'A{row}'].font = header_font

        row += 1
        worksheet[f'A{row}'] = f"Total Items: {totals.get('item_count', 0)}"
        worksheet[f'D{row}'] = f"Total Quantity: {totals.get('total_quantity', 0)}"

        row += 1
        worksheet[f'A{row}'] = f"Total Value: {totals.get('total_value', 0)}"
        worksheet[f'D{row}'] = f"Total Amount: {totals.get('total_line_amount', 0)}"

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        return workbook
